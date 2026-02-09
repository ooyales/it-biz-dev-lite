#!/usr/bin/env python3
"""
Pricing & Budget Generator Agent
Creates pricing models and IGCEs (Independent Government Cost Estimates)

Features:
- Labor rate calculations
- Loaded rates (burden, overhead, G&A, profit)
- IGCE generation
- Price volume creation
- Cost comparison analysis
"""

import sys
import os
from typing import Dict, List
from datetime import datetime
from pathlib import Path

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("⚠️  openpyxl not installed")


class LaborCategory:
    """Represents a labor category with rates"""
    
    def __init__(self, name: str, base_rate: float, clearance_req: str = None):
        self.name = name
        self.base_rate = base_rate  # $/hour
        self.clearance_req = clearance_req
        
        # Burden rates (customize for your company)
        self.fringe_rate = 0.30  # 30% fringe benefits
        self.overhead_rate = 0.45  # 45% overhead
        self.ga_rate = 0.08  # 8% G&A
        self.profit_rate = 0.10  # 10% profit/fee
        
    @property
    def loaded_rate(self) -> float:
        """Calculate fully loaded hourly rate"""
        
        # Base + Fringe
        base_plus_fringe = self.base_rate * (1 + self.fringe_rate)
        
        # + Overhead
        with_overhead = base_plus_fringe * (1 + self.overhead_rate)
        
        # + G&A
        with_ga = with_overhead * (1 + self.ga_rate)
        
        # + Profit
        fully_loaded = with_ga * (1 + self.profit_rate)
        
        return round(fully_loaded, 2)
    
    @property
    def annual_cost(self) -> float:
        """Annual cost at 2080 hours/year"""
        return self.loaded_rate * 2080


class PricingModel:
    """Generates pricing models and IGCEs"""
    
    def __init__(self):
        # Standard labor categories (customize for your company)
        self.labor_categories = [
            LaborCategory('Program Manager', 95.00, 'Secret'),
            LaborCategory('Senior Software Engineer', 85.00, 'Secret'),
            LaborCategory('Software Engineer', 65.00, 'Secret'),
            LaborCategory('Junior Software Engineer', 45.00),
            LaborCategory('Senior Cybersecurity Analyst', 90.00, 'Top Secret'),
            LaborCategory('Cybersecurity Analyst', 70.00, 'Secret'),
            LaborCategory('Data Scientist', 80.00),
            LaborCategory('Cloud Architect', 88.00, 'Secret'),
            LaborCategory('DevOps Engineer', 75.00),
            LaborCategory('Business Analyst', 60.00),
            LaborCategory('Technical Writer', 55.00),
            LaborCategory('Administrative Support', 40.00),
        ]
        
        # Other Direct Costs
        self.odc_categories = {
            'Travel': 0,
            'Materials': 0,
            'Equipment': 0,
            'Subcontractors': 0,
            'Other': 0
        }
    
    def calculate_labor_cost(self, staffing: Dict[str, float], 
                           duration_months: int = 12) -> Dict:
        """Calculate labor costs for staffing plan"""
        
        results = {
            'labor_categories': [],
            'total_hours': 0,
            'total_cost': 0
        }
        
        # Hours per month (average)
        hours_per_month = 173.33  # 2080 hours/year / 12 months
        
        for cat_name, fte_count in staffing.items():
            # Find labor category
            cat = next((c for c in self.labor_categories if c.name == cat_name), None)
            if not cat:
                print(f"⚠️  Unknown labor category: {cat_name}")
                continue
            
            # Calculate hours and cost
            total_hours = fte_count * hours_per_month * duration_months
            total_cost = total_hours * cat.loaded_rate
            
            results['labor_categories'].append({
                'name': cat_name,
                'fte': fte_count,
                'hours': total_hours,
                'rate': cat.loaded_rate,
                'cost': total_cost,
                'clearance': cat.clearance_req
            })
            
            results['total_hours'] += total_hours
            results['total_cost'] += total_cost
        
        return results
    
    def generate_igce(self, opportunity: Dict, staffing: Dict[str, float], 
                     duration_months: int = 12, odc: Dict = None) -> Dict:
        """Generate Independent Government Cost Estimate"""
        
        # Calculate labor
        labor_results = self.calculate_labor_cost(staffing, duration_months)
        
        # Add ODCs
        odc_total = 0
        if odc:
            for category, amount in odc.items():
                odc_total += amount
        
        # Total contract value
        total_value = labor_results['total_cost'] + odc_total
        
        igce = {
            'opportunity': opportunity.get('title', 'Unknown'),
            'agency': opportunity.get('agency', 'Unknown'),
            'duration_months': duration_months,
            'labor': labor_results,
            'odc': odc or {},
            'odc_total': odc_total,
            'total_value': total_value,
            'monthly_burn': total_value / duration_months,
            'generated_date': datetime.now().strftime('%B %d, %Y')
        }
        
        return igce
    
    def generate_pricing_excel(self, igce: Dict, filename: str = None) -> str:
        """Generate Excel pricing workbook"""
        
        if not OPENPYXL_AVAILABLE:
            return self.generate_pricing_text(igce)
        
        wb = Workbook()
        
        # Summary sheet
        summary = wb.active
        summary.title = 'Summary'
        
        # Title
        summary['A1'] = 'PRICING SUMMARY'
        summary['A1'].font = Font(size=14, bold=True)
        
        # Opportunity info
        summary['A3'] = 'Opportunity:'
        summary['A3'].font = Font(bold=True)
        summary['B3'] = igce['opportunity']
        
        summary['A4'] = 'Agency:'
        summary['A4'].font = Font(bold=True)
        summary['B4'] = igce['agency']
        
        summary['A5'] = 'Period of Performance:'
        summary['A5'].font = Font(bold=True)
        summary['B5'] = f"{igce['duration_months']} months"
        
        # Cost summary
        summary['A7'] = 'COST SUMMARY'
        summary['A7'].font = Font(size=12, bold=True)
        
        summary['A8'] = 'Labor Cost:'
        summary['B8'] = igce['labor']['total_cost']
        summary['B8'].number_format = '$#,##0.00'
        
        summary['A9'] = 'ODC Cost:'
        summary['B9'] = igce['odc_total']
        summary['B9'].number_format = '$#,##0.00'
        
        summary['A10'] = 'Total Contract Value:'
        summary['A10'].font = Font(bold=True)
        summary['B10'] = '=B8+B9'
        summary['B10'].number_format = '$#,##0.00'
        summary['B10'].font = Font(bold=True, color='0000FF')
        
        summary['A11'] = 'Monthly Burn Rate:'
        summary['B11'] = f'=B10/{igce["duration_months"]}'
        summary['B11'].number_format = '$#,##0.00'
        
        # Labor detail sheet
        labor = wb.create_sheet('Labor Rates')
        
        # Headers
        headers = ['Labor Category', 'FTE', 'Hours', 'Loaded Rate', 'Total Cost', 'Clearance']
        for col, header in enumerate(headers, 1):
            cell = labor.cell(1, col)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', start_color='4472C4')
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        row = 2
        for cat in igce['labor']['labor_categories']:
            labor.cell(row, 1).value = cat['name']
            labor.cell(row, 2).value = cat['fte']
            labor.cell(row, 2).number_format = '0.00'
            labor.cell(row, 3).value = cat['hours']
            labor.cell(row, 3).number_format = '#,##0'
            labor.cell(row, 4).value = cat['rate']
            labor.cell(row, 4).number_format = '$#,##0.00'
            labor.cell(row, 5).value = cat['cost']
            labor.cell(row, 5).number_format = '$#,##0.00'
            labor.cell(row, 6).value = cat['clearance'] or 'None'
            row += 1
        
        # Totals
        labor.cell(row, 1).value = 'TOTAL'
        labor.cell(row, 1).font = Font(bold=True)
        labor.cell(row, 2).value = f'=SUM(B2:B{row-1})'
        labor.cell(row, 2).font = Font(bold=True)
        labor.cell(row, 3).value = f'=SUM(C2:C{row-1})'
        labor.cell(row, 3).font = Font(bold=True)
        labor.cell(row, 5).value = f'=SUM(E2:E{row-1})'
        labor.cell(row, 5).font = Font(bold=True, color='0000FF')
        labor.cell(row, 5).number_format = '$#,##0.00'
        
        # Column widths
        labor.column_dimensions['A'].width = 30
        labor.column_dimensions['B'].width = 10
        labor.column_dimensions['C'].width = 12
        labor.column_dimensions['D'].width = 15
        labor.column_dimensions['E'].width = 15
        labor.column_dimensions['F'].width = 15
        
        # ODC sheet
        odc_sheet = wb.create_sheet('Other Direct Costs')
        
        odc_sheet['A1'] = 'Category'
        odc_sheet['A1'].font = Font(bold=True)
        odc_sheet['B1'] = 'Cost'
        odc_sheet['B1'].font = Font(bold=True)
        
        row = 2
        for category, cost in igce['odc'].items():
            odc_sheet.cell(row, 1).value = category
            odc_sheet.cell(row, 2).value = cost
            odc_sheet.cell(row, 2).number_format = '$#,##0.00'
            row += 1
        
        # Total
        odc_sheet.cell(row, 1).value = 'TOTAL ODC'
        odc_sheet.cell(row, 1).font = Font(bold=True)
        odc_sheet.cell(row, 2).value = f'=SUM(B2:B{row-1})'
        odc_sheet.cell(row, 2).font = Font(bold=True, color='0000FF')
        odc_sheet.cell(row, 2).number_format = '$#,##0.00'
        
        # Save
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"pricing_{timestamp}.xlsx"
        
        wb.save(filename)
        return filename
    
    def generate_pricing_text(self, igce: Dict) -> str:
        """Generate text pricing as fallback"""
        
        lines = []
        lines.append("="*70)
        lines.append("PRICING SUMMARY")
        lines.append("="*70)
        lines.append(f"\nOpportunity: {igce['opportunity']}")
        lines.append(f"Agency: {igce['agency']}")
        lines.append(f"Period: {igce['duration_months']} months")
        lines.append(f"Generated: {igce['generated_date']}")
        lines.append("\n" + "-"*70)
        lines.append("COST SUMMARY")
        lines.append("-"*70)
        lines.append(f"Labor Cost:              ${igce['labor']['total_cost']:,.2f}")
        lines.append(f"ODC Cost:                ${igce['odc_total']:,.2f}")
        lines.append(f"Total Contract Value:    ${igce['total_value']:,.2f}")
        lines.append(f"Monthly Burn Rate:       ${igce['monthly_burn']:,.2f}")
        
        lines.append("\n" + "-"*70)
        lines.append("LABOR BREAKDOWN")
        lines.append("-"*70)
        lines.append(f"{'Category':<30} {'FTE':>6} {'Hours':>10} {'Rate':>12} {'Cost':>12}")
        lines.append("-"*70)
        
        for cat in igce['labor']['labor_categories']:
            lines.append(f"{cat['name']:<30} {cat['fte']:>6.2f} {cat['hours']:>10.0f} "
                        f"${cat['rate']:>11.2f} ${cat['cost']:>11,.2f}")
        
        lines.append("-"*70)
        lines.append(f"{'TOTAL':<30} {'':<6} {igce['labor']['total_hours']:>10.0f} "
                    f"{'':<12} ${igce['labor']['total_cost']:>11,.2f}")
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"pricing_{timestamp}.txt"
        
        with open(filename, 'w') as f:
            f.write('\n'.join(lines))
        
        return filename


def main():
    """Test pricing generator"""
    
    print("\n💰 PRICING & BUDGET GENERATOR")
    print("="*70)
    
    # Test opportunity
    opportunity = {
        'title': 'Cloud Migration Services',
        'agency': 'Department of Defense'
    }
    
    # Staffing plan (FTEs)
    staffing = {
        'Program Manager': 1.0,
        'Senior Software Engineer': 2.0,
        'Software Engineer': 4.0,
        'Cloud Architect': 1.0,
        'DevOps Engineer': 2.0,
        'Cybersecurity Analyst': 1.0
    }
    
    # ODCs
    odc = {
        'Travel': 50000,
        'Training': 25000,
        'Software Licenses': 75000,
        'Hardware': 100000
    }
    
    pricing = PricingModel()
    
    print(f"Generating pricing for: {opportunity['title']}")
    print(f"Agency: {opportunity['agency']}")
    print(f"Duration: 12 months")
    print()
    
    # Generate IGCE
    igce = pricing.generate_igce(opportunity, staffing, duration_months=12, odc=odc)
    
    print(f"Labor Cost: ${igce['labor']['total_cost']:,.2f}")
    print(f"ODC Cost: ${igce['odc_total']:,.2f}")
    print(f"Total Value: ${igce['total_value']:,.2f}")
    print(f"Monthly Burn: ${igce['monthly_burn']:,.2f}")
    print()
    
    # Generate Excel
    filename = pricing.generate_pricing_excel(igce)
    
    print(f"✓ Pricing workbook saved: {filename}")
    print()


if __name__ == '__main__':
    main()
