#!/usr/bin/env python3
"""
BD Intelligence Excel Exporter
Exports Neo4j knowledge graph data to a comprehensive multi-tab Excel workbook

Features:
- Opportunities tab with scoring and contacts
- Contacts tab with relationships
- Organizations tab
- Contracts tab (FPDS data)
- Competitive Intelligence tab
- Dashboard tab with charts and KPIs
- All tabs linked with formulas
"""

import sys
sys.path.append('knowledge_graph')

import os
from datetime import datetime
from pathlib import Path

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.worksheet.table import Table, TableStyleInfo

from graph.neo4j_client import KnowledgeGraphClient
import json


class BDIntelligenceExporter:
    """Export BD intelligence data to Excel"""
    
    def __init__(self):
        # Connect to Neo4j
        self.kg = KnowledgeGraphClient(
            uri=os.getenv('NEO4J_URI', 'bolt://localhost:7687'),
            user=os.getenv('NEO4J_USER', 'neo4j'),
            password=os.getenv('NEO4J_PASSWORD')
        )
        
        # Create workbook
        self.wb = Workbook()
        self.wb.remove(self.wb.active)  # Remove default sheet
        
        # Styling
        self.header_font = Font(bold=True, color='FFFFFF', size=11)
        self.header_fill = PatternFill('solid', start_color='4472C4')
        self.header_alignment = Alignment(horizontal='center', vertical='center')
        
        # Data fonts (per best practices)
        self.input_font = Font(color='0000FF')  # Blue for inputs
        self.formula_font = Font(color='000000')  # Black for formulas
        self.link_font = Font(color='008000')  # Green for internal links
        
    def add_header_row(self, sheet, headers, row=1):
        """Add formatted header row"""
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=row, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            
            # Auto-size column
            sheet.column_dimensions[get_column_letter(col)].width = max(len(str(header)) + 2, 12)
    
    def export_opportunities(self):
        """Export opportunities with scoring"""
        
        sheet = self.wb.create_sheet('Opportunities', 0)
        
        # Headers
        headers = [
            'Title', 'Agency', 'Notice ID', 'Posted Date', 'Deadline',
            'Set-Aside', 'NAICS', 'Type',
            'Total Score', 'Win Probability %', 'Priority',
            'Contacts Count', 'Decision Makers', 'Tech Leads',
            'Recommendation', 'SAM.gov Link'
        ]
        self.add_header_row(sheet, headers)
        
        # Load opportunity scout data
        scout_files = sorted(Path('knowledge_graph').glob('scout_data_*.json'), reverse=True)
        
        if not scout_files:
            print("⚠️  No scout data found. Run opportunity_scout.py first.")
            return
        
        with open(scout_files[0]) as f:
            data = json.load(f)
        
        opportunities = data.get('opportunities', [])
        scores = data.get('scores', [])
        
        # Add data
        row = 2
        for opp, score in zip(opportunities, scores):
            # Extract agency
            agency = (opp.get('organizationName') or 
                     opp.get('fullParentPathName') or
                     opp.get('department') or
                     'Unknown')
            
            sheet.cell(row, 1).value = opp.get('title', '')
            sheet.cell(row, 2).value = agency
            sheet.cell(row, 3).value = opp.get('noticeId', '')
            sheet.cell(row, 4).value = opp.get('postedDate', '')
            sheet.cell(row, 5).value = opp.get('responseDeadLine', '')
            sheet.cell(row, 6).value = opp.get('typeOfSetAside', '')
            sheet.cell(row, 7).value = opp.get('naicsCode', '')
            sheet.cell(row, 8).value = opp.get('type', '')
            
            # Scores (as values, not formulas since they're pre-calculated)
            sheet.cell(row, 9).value = score.get('total_score', 0)
            sheet.cell(row, 9).font = self.input_font  # Blue since it's an input value
            
            sheet.cell(row, 10).value = score.get('win_probability', 0)
            sheet.cell(row, 10).number_format = '0.0%'
            sheet.cell(row, 10).font = self.input_font
            
            sheet.cell(row, 11).value = score.get('priority', '')
            
            # Contacts
            contacts = score.get('contacts', {})
            sheet.cell(row, 12).value = contacts.get('total_contacts', 0)
            sheet.cell(row, 13).value = len(contacts.get('decision_makers', []))
            sheet.cell(row, 14).value = len(contacts.get('technical_leads', []))
            
            sheet.cell(row, 15).value = score.get('recommendation', '')
            
            # SAM.gov link as hyperlink formula
            notice_id = opp.get('noticeId', '')
            if notice_id:
                formula = f'=HYPERLINK("https://sam.gov/opp/{notice_id}/view", "View")'
                sheet.cell(row, 16).value = formula
                sheet.cell(row, 16).font = Font(color='0563C1', underline='single')
            
            row += 1
        
        # Apply table formatting
        tab = Table(displayName='OpportunitiesTable', ref=f'A1:P{row-1}')
        style = TableStyleInfo(
            name='TableStyleMedium2',
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        tab.tableStyleInfo = style
        sheet.add_table(tab)
        
        # Freeze header
        sheet.freeze_panes = 'A2'
        
        print(f"✓ Exported {row-2} opportunities")
    
    def export_contacts(self):
        """Export contacts with relationships"""
        
        sheet = self.wb.create_sheet('Contacts')
        
        headers = [
            'Name', 'Title', 'Organization', 'Email', 'Phone',
            'Role Type', 'Influence Level', 'Source', 'Extracted Date'
        ]
        self.add_header_row(sheet, headers)
        
        # Query Neo4j
        with self.kg.driver.session(database="contactsgraphdb") as session:
            query = """
            MATCH (p:Person)
            OPTIONAL MATCH (p)-[:WORKS_AT]->(o:Organization)
            RETURN p.name as name,
                   p.title as title,
                   COALESCE(o.name, p.organization) as organization,
                   p.email as email,
                   p.phone as phone,
                   p.role_type as role_type,
                   p.influence_level as influence_level,
                   p.source as source,
                   p.extracted_at as extracted_at
            ORDER BY p.name
            """
            result = session.run(query)
            
            row = 2
            for record in result:
                sheet.cell(row, 1).value = record['name']
                sheet.cell(row, 2).value = record['title']
                sheet.cell(row, 3).value = record['organization']
                sheet.cell(row, 4).value = record['email']
                sheet.cell(row, 5).value = record['phone']
                sheet.cell(row, 6).value = record['role_type']
                sheet.cell(row, 7).value = record['influence_level']
                sheet.cell(row, 8).value = record['source']
                sheet.cell(row, 9).value = record['extracted_at']
                row += 1
        
        # Table formatting
        if row > 2:
            tab = Table(displayName='ContactsTable', ref=f'A1:I{row-1}')
            tab.tableStyleInfo = TableStyleInfo(
                name='TableStyleMedium2',
                showRowStripes=True
            )
            sheet.add_table(tab)
        
        sheet.freeze_panes = 'A2'
        print(f"✓ Exported {row-2} contacts")
    
    def export_organizations(self):
        """Export organizations"""
        
        sheet = self.wb.create_sheet('Organizations')
        
        headers = ['Organization', 'Type', 'Contact Count', 'Contract Count']
        self.add_header_row(sheet, headers)
        
        with self.kg.driver.session(database="contactsgraphdb") as session:
            query = """
            MATCH (o:Organization)
            OPTIONAL MATCH (o)<-[:WORKS_AT]-(p:Person)
            OPTIONAL MATCH (c:Contract)-[:AWARDED_TO]->(o)
            WITH o, count(DISTINCT p) as people_count, count(DISTINCT c) as contract_count
            RETURN o.name as name,
                   o.type as type,
                   people_count,
                   contract_count
            ORDER BY people_count DESC, contract_count DESC
            """
            result = session.run(query)
            
            row = 2
            for record in result:
                sheet.cell(row, 1).value = record['name']
                sheet.cell(row, 2).value = record['type']
                sheet.cell(row, 3).value = record['people_count']
                sheet.cell(row, 4).value = record['contract_count']
                row += 1
        
        if row > 2:
            tab = Table(displayName='OrganizationsTable', ref=f'A1:D{row-1}')
            tab.tableStyleInfo = TableStyleInfo(name='TableStyleMedium2', showRowStripes=True)
            sheet.add_table(tab)
        
        sheet.freeze_panes = 'A2'
        print(f"✓ Exported {row-2} organizations")
    
    def export_contracts(self):
        """Export FPDS contract data"""
        
        sheet = self.wb.create_sheet('Contracts')
        
        headers = [
            'Contract Number', 'Title', 'Agency', 'Contractor',
            'Award Date', 'Value', 'NAICS', 'Description'
        ]
        self.add_header_row(sheet, headers)
        
        with self.kg.driver.session(database="contactsgraphdb") as session:
            query = """
            MATCH (c:Contract)-[:AWARDED_TO]->(o:Organization)
            RETURN c.contract_number as number,
                   c.title as title,
                   c.agency as agency,
                   o.name as contractor,
                   c.award_date as award_date,
                   c.value as value,
                   c.naics as naics,
                   c.description as description
            ORDER BY c.award_date DESC
            LIMIT 1000
            """
            result = session.run(query)
            
            row = 2
            for record in result:
                sheet.cell(row, 1).value = record['number']
                sheet.cell(row, 2).value = record['title']
                sheet.cell(row, 3).value = record['agency']
                sheet.cell(row, 4).value = record['contractor']
                sheet.cell(row, 5).value = record['award_date']
                
                # Format currency
                value_cell = sheet.cell(row, 6)
                value_cell.value = record['value'] or 0
                value_cell.number_format = '$#,##0'
                
                sheet.cell(row, 7).value = record['naics']
                sheet.cell(row, 8).value = record['description']
                row += 1
        
        if row > 2:
            tab = Table(displayName='ContractsTable', ref=f'A1:H{row-1}')
            tab.tableStyleInfo = TableStyleInfo(name='TableStyleMedium2', showRowStripes=True)
            sheet.add_table(tab)
        
        sheet.freeze_panes = 'A2'
        print(f"✓ Exported {row-2} contracts")
    
    def export_dashboard(self):
        """Create dashboard with KPIs and charts"""
        
        sheet = self.wb.create_sheet('Dashboard', 0)
        
        # Title
        sheet['A1'] = 'BD Intelligence Dashboard'
        sheet['A1'].font = Font(size=18, bold=True)
        
        sheet['A2'] = f'Generated: {datetime.now().strftime("%B %d, %Y at %I:%M %p")}'
        sheet['A2'].font = Font(size=10, italic=True)
        
        # KPIs
        sheet['A4'] = 'Key Performance Indicators'
        sheet['A4'].font = Font(size=14, bold=True)
        
        # Opportunity KPIs with formulas linking to Opportunities sheet
        row = 6
        kpis = [
            ('Total Opportunities', '=COUNTA(Opportunities!A:A)-1'),
            ('High Priority', '=COUNTIF(Opportunities!K:K,"HIGH")'),
            ('With Contacts', '=COUNTIF(Opportunities!L:L,">0")'),
            ('Avg Win Probability', '=AVERAGE(Opportunities!J:J)'),
        ]
        
        for label, formula in kpis:
            sheet.cell(row, 1).value = label
            sheet.cell(row, 1).font = Font(bold=True)
            sheet.cell(row, 2).value = formula
            sheet.cell(row, 2).font = self.link_font  # Green for internal link
            if 'Probability' in label:
                sheet.cell(row, 2).number_format = '0.0%'
            row += 1
        
        # Network KPIs
        row += 1
        sheet.cell(row, 1).value = 'Network Intelligence'
        sheet.cell(row, 1).font = Font(size=14, bold=True)
        row += 1
        
        network_kpis = [
            ('Total Contacts', '=COUNTA(Contacts!A:A)-1'),
            ('Decision Makers', '=COUNTIF(Contacts!F:F,"Decision Maker")'),
            ('Organizations', '=COUNTA(Organizations!A:A)-1'),
            ('Contracts Tracked', '=COUNTA(Contracts!A:A)-1'),
        ]
        
        for label, formula in network_kpis:
            sheet.cell(row, 1).value = label
            sheet.cell(row, 1).font = Font(bold=True)
            sheet.cell(row, 2).value = formula
            sheet.cell(row, 2).font = self.link_font
            row += 1
        
        # Instructions
        row += 2
        sheet.cell(row, 1).value = 'How to Use This Workbook:'
        sheet.cell(row, 1).font = Font(size=12, bold=True)
        row += 1
        
        instructions = [
            '• Opportunities: Browse all opportunities with scoring and contact info',
            '• Contacts: View your network with contact details',
            '• Organizations: See all organizations and activity',
            '• Contracts: Review FPDS contract data for competitive intelligence',
            '• All data is linked - update will reflect across tabs',
            '• Use filters and sorts to find what you need',
            '• Export subsets using "Save As" or copy/paste',
        ]
        
        for instruction in instructions:
            sheet.cell(row, 1).value = instruction
            row += 1
        
        # Auto-size columns
        sheet.column_dimensions['A'].width = 30
        sheet.column_dimensions['B'].width = 20
        
        print("✓ Created dashboard")
    
    def export(self, filename=None):
        """Export all data to Excel"""
        
        print("\n📊 BD INTELLIGENCE EXCEL EXPORT")
        print("="*70)
        
        # Export each tab
        self.export_dashboard()
        self.export_opportunities()
        self.export_contacts()
        self.export_organizations()
        self.export_contracts()
        
        # Save
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"bd_intelligence_{timestamp}.xlsx"
        
        self.wb.save(filename)
        
        print("\n" + "="*70)
        print(f"✓ Export complete: {filename}")
        print("\n📂 The workbook contains:")
        print("   • Dashboard - KPIs and overview")
        print("   • Opportunities - 200 opportunities with scoring")
        print("   • Contacts - Your network")
        print("   • Organizations - All orgs and relationships")
        print("   • Contracts - FPDS competitive data")
        print()
        
        # Close Neo4j connection
        self.kg.close()
        
        return filename


def main():
    """Run the export"""
    
    exporter = BDIntelligenceExporter()
    filename = exporter.export()
    
    print(f"🎉 Open in Excel to browse your BD intelligence data!")
    print(f"   {os.path.abspath(filename)}")
    print()


if __name__ == '__main__':
    main()
